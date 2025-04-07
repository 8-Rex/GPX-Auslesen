import os
from xml.etree import ElementTree as ET
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox

def parse_gpx(file_path):
    tree = ET.parse(file_path)
    root = tree.getroot()
    namespace = {'default': 'http://www.topografix.com/GPX/1/1'}

    # Versuche den Namen aus dem Metadata-Tag
    metadata_name = root.find("default:metadata/default:name", namespace)
    if metadata_name is not None:
        full_name = metadata_name.text.split(" - ")[0].strip()
        name_parts = full_name.split()

        if len(name_parts) >= 2:
            last_name_candidates = []
            while name_parts and ('-' in name_parts[-1] or not last_name_candidates):
                last_name_candidates.insert(0, name_parts.pop())
            first_name = " ".join(name_parts)
            last_name = " ".join(last_name_candidates)
        else:
            first_name = full_name
            last_name = ""
    else:
        file_basename = os.path.basename(file_path)
        name_candidate = file_basename.split("_")[-1].split(".")[0]
        name_parts = name_candidate.split()

        if len(name_parts) >= 2:
            last_name_candidates = []
            while name_parts and ('-' in name_parts[-1] or not last_name_candidates):
                last_name_candidates.insert(0, name_parts.pop())
            first_name = " ".join(name_parts)
            last_name = " ".join(last_name_candidates)
        else:
            first_name = name_candidate
            last_name = ""

    # Runde extrahieren
    track_name_elem = root.find(".//default:trk/default:name", namespace)
    if track_name_elem is not None:
        track_name = track_name_elem.text
        if "Runde" in track_name:
            parts = track_name.split("Runde")
            if len(parts) > 1:
                num = ''.join(filter(str.isdigit, parts[1]))
                track_number = f"Runde {num}" if num else "Unbekannt"
            else:
                track_number = "Unbekannt"
        else:
            track_number = "Unbekannt"
    else:
        track_number = "Unbekannt"

    # Zeitstempel
    timestamps = []
    for trkpt in root.findall(".//default:trkpt", namespace):
        time_elem = trkpt.find("default:time", namespace)
        if time_elem is not None:
            timestamps.append(time_elem.text)

    if timestamps:
        times = [datetime.fromisoformat(ts.replace("Z", "+00:00")) for ts in timestamps]
        total_duration = times[-1] - times[0]

        movement_duration = timedelta()
        pause_threshold = 10  # Sekunden
        for i in range(1, len(times)):
            delta = times[i] - times[i - 1]
            if delta.total_seconds() <= pause_threshold:
                movement_duration += delta
    else:
        total_duration = timedelta()
        movement_duration = timedelta()

    return first_name, last_name, track_number, total_duration, movement_duration

def format_timedelta(td):
    total_seconds = int(td.total_seconds())
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{hours:02}:{minutes:02}:{seconds:02}"

def save_to_excel(results, excel_path):
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
        existing_data = set(tuple(row) for row in ws.iter_rows(values_only=True))
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Vorname", "Nachname", "Runde", "Gesamtdauer (hh:mm:ss)", "Bewegungszeit (hh:mm:ss)"])
        existing_data = set()

    new_entries = 0
    for result in results:
        if tuple(result) not in existing_data:
            ws.append(result)
            new_entries += 1

    wb.save(excel_path)
    return new_entries

def show_preview(results, excel_path):
    preview_win = tk.Toplevel()
    preview_win.title("Vorschau der Ergebnisse")
    preview_win.geometry("800x500")

    text = tk.Text(preview_win, wrap=tk.NONE, font=("Courier", 10))
    text.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)

    header = f"{'Vorname':<20} {'Nachname':<25} {'Runde':<15} {'Gesamtdauer':<15} {'Bewegungszeit':<15}\n"
    text.insert(tk.END, header)
    text.insert(tk.END, "-" * 95 + "\n")

    for row in results:
        line = f"{row[0]:<20} {row[1]:<25} {row[2]:<15} {row[3]:<15} {row[4]:<15}\n"
        text.insert(tk.END, line)

    text.config(state=tk.DISABLED)

    # Button-Leiste
    btn_frame = tk.Frame(preview_win)
    btn_frame.pack(pady=10)

    def speichern():
        new_entries = save_to_excel(results, excel_path)
        messagebox.showinfo("Fertig", f"{new_entries} neue Einträge wurden gespeichert.")
        preview_win.destroy()

    def abbrechen():
        preview_win.destroy()

    btn_speichern = tk.Button(btn_frame, text="In Excel speichern", command=speichern)
    btn_abbrechen = tk.Button(btn_frame, text="Abbrechen", command=abbrechen)
    btn_speichern.pack(side=tk.LEFT, padx=10)
    btn_abbrechen.pack(side=tk.LEFT, padx=10)

def process_gpx_files(gpx_folder, excel_path):
    gpx_files = [f for f in os.listdir(gpx_folder) if f.lower().endswith(".gpx")]
    if not gpx_files:
        messagebox.showinfo("Keine Dateien", "Im ausgewählten Ordner wurden keine GPX-Dateien gefunden.")
        return

    results = []
    for gpx_file in gpx_files:
        file_path = os.path.join(gpx_folder, gpx_file)
        first_name, last_name, track_number, total_duration, movement_duration = parse_gpx(file_path)
        result = [
            first_name,
            last_name,
            track_number,
            format_timedelta(total_duration),
            format_timedelta(movement_duration)
        ]
        results.append(result)
        print("[DEBUG]", result)  # Debug-Ausgabe

    show_preview(results, excel_path)

def start_processing():
    gpx_folder = filedialog.askdirectory(title="Ordner mit GPX-Dateien auswählen")
    if not gpx_folder:
        return

    excel_path = filedialog.asksaveasfilename(
        title="Excel-Datei speichern unter...",
        defaultextension=".xlsx",
        filetypes=[("Excel-Dateien", "*.xlsx")]
    )
    if not excel_path:
        return

    process_gpx_files(gpx_folder, excel_path)

# GUI Setup
if __name__ == "__main__":
    root = tk.Tk()
    root.title("GPX-Auswertung")
    root.geometry("300x150")

    label = tk.Label(root, text="GPX-Auswertung in Excel", font=("Arial", 12))
    label.pack(pady=20)

    button = tk.Button(root, text="GPX-Dateien verarbeiten", command=start_processing)
    button.pack(pady=10)

    root.mainloop()
